import io
from datetime import UTC, datetime, timedelta

import pytest
from openpyxl import load_workbook

from farmers_chatbot.artifacts import ArtifactService
from farmers_chatbot.auth import UserIdentity
from farmers_chatbot.pilot_store import PilotStore
from farmers_chatbot.storage_backends import LocalPrivateStorage
from farmers_chatbot.tools import ToolRegistry
from farmers_chatbot.trusted_sources import (
    LiveEvidence,
    LiveSourceDefinition,
    TrustedSourceClient,
    assess_claim_support,
    live_evidence_is_stale,
)


class _DirectResponse:
    status_code = 200

    def __init__(
        self,
        passage: str,
        *,
        status_code: int = 200,
        content_type: str = "text/plain",
    ) -> None:
        self.status_code = status_code
        self.payload = passage.encode()
        self.headers = {
            "Content-Type": content_type,
            "Content-Length": str(len(self.payload)),
        }

    def iter_content(self, chunk_size: int):
        for index in range(0, len(self.payload), chunk_size):
            yield self.payload[index : index + chunk_size]


class _DirectSession:
    def __init__(self, response: _DirectResponse) -> None:
        self.response = response
        self.calls = []

    def get(self, url, **kwargs):
        self.calls.append((url, kwargs))
        return self.response


def _definition(**overrides) -> LiveSourceDefinition:
    values = {
        "source_id": "FAO-MARKET-BULLETIN",
        "publisher": "Food and Agriculture Organization",
        "title": "FAO potato market bulletin",
        "url": "https://www.fao.org/markets/potato",
        "categories": ("economic",),
        "ttl_seconds": 3600,
        "authorized": True,
        "keywords": ("potato", "price", "market"),
    }
    values.update(overrides)
    return LiveSourceDefinition(**values)


def _identity() -> UserIdentity:
    return UserIdentity(
        user_id="",
        issuer="https://accounts.google.com",
        subject="artifact-user",
        email="artifact@example.org",
        name="Artifact Tester",
        is_admin=False,
    )


def _artifact_service(tmp_path):
    store = PilotStore(sqlite_path=tmp_path / "pilot.sqlite3")
    storage = LocalPrivateStorage(tmp_path / "private")
    user = store.upsert_user(_identity())
    return store, storage, user, ArtifactService(
        store,
        storage,
        owner_user_id=user["id"],
    )


def test_live_search_retains_direct_source_passage() -> None:
    passage = "The reported potato price is 40 USD per unit for the observed market."
    session = _DirectSession(_DirectResponse(passage))
    result = TrustedSourceClient(
        "unused-compatibility-key",
        enabled=True,
        definitions=(_definition(),),
        session=session,  # type: ignore[arg-type]
    ).search(
        "What is the latest potato price?",
        "economic",
    )

    assert result.verified
    assert result.support_status == "supported"
    assert result.summary == passage
    assert len(result.citations) == 1
    evidence = result.evidence[0]
    assert evidence.evidence_id.startswith("live_")
    assert evidence.publisher == "Food and Agriculture Organization"
    assert evidence.passage == passage
    assert evidence.support_status == "supported"
    assert evidence.live_only and not evidence.stale
    assert datetime.fromisoformat(evidence.expires_at) > datetime.fromisoformat(
        evidence.observed_at
    )
    assert session.calls[0][1]["allow_redirects"] is False
    assert session.calls[0][1]["stream"] is True


def test_authorized_endpoint_alone_is_not_relevant_evidence() -> None:
    result = TrustedSourceClient(
        None,
        enabled=True,
        definitions=(_definition(keywords=()),),
        session=_DirectSession(
            _DirectResponse("Rainfall observations were recorded in another region.")
        ),  # type: ignore[arg-type]
    ).search(
        "What is the latest potato price?",
        "economic",
    )

    assert not result.verified
    assert result.support_status == "unassessed"
    assert result.summary == ""
    assert result.citations == ()
    assert result.evidence == ()
    assert "no relevant" in (result.warning or "")


@pytest.mark.parametrize(
    ("response", "max_bytes"),
    [
        (_DirectResponse("potato price bulletin", status_code=302), 1_000_000),
        (_DirectResponse("potato price bulletin", content_type="application/pdf"), 1_000_000),
        (_DirectResponse("potato price " * 1000), 1024),
    ],
)
def test_redirect_content_type_and_size_fail_closed(response, max_bytes) -> None:
    result = TrustedSourceClient(
        None,
        enabled=True,
        definitions=(_definition(max_bytes=max_bytes),),
        session=_DirectSession(response),  # type: ignore[arg-type]
    ).search(
        "latest potato price",
        "economic",
    )

    assert not result.verified
    assert result.support_status == "unassessed"
    assert result.citations == ()


def test_live_evidence_staleness_and_negation_are_fail_closed() -> None:
    now = datetime.now(UTC)
    assert live_evidence_is_stale((now - timedelta(seconds=1)).isoformat(), now=now)
    assert not live_evidence_is_stale(
        (now + timedelta(seconds=1)).isoformat(), now=now
    )
    assert live_evidence_is_stale("not-a-date", now=now)
    assert (
        assess_claim_support("This pesticide is not allowed", "This pesticide is allowed")
        == "unsupported"
    )
    assert (
        assess_claim_support(
            "The potato price is 40 USD",
            "The potato price is 50 USD",
        )
        == "unsupported"
    )
    stale = LiveEvidence(
        evidence_id="live_old",
        publisher="FAO",
        title="Old bulletin",
        url="https://www.fao.org/old",
        passage="Old passage",
        claim="Old claim",
        observed_at=(now - timedelta(days=2)).isoformat(),
        expires_at=(now - timedelta(days=1)).isoformat(),
        category="economic",
        support_status="supported",
    )
    assert stale.stale and not stale.usable


def test_enterprise_budget_builds_decision_sheets_and_provenance(tmp_path) -> None:
    store, storage, user, service = _artifact_service(tmp_path)
    result = service.calculate_enterprise_budget(
        title="Potato enterprise",
        currency="USD",
        geography="Akkar, Lebanon",
        as_of_date="2026-08-11",
        costs=[
            {
                "item": "Seed",
                "quantity": 2,
                "unit": "bag",
                "unit_cost": 100,
                "period": "2026-01",
                "value_status": "sourced",
                "evidence_id": "live_seed",
            }
        ],
        revenues=[
            {
                "item": "Potatoes",
                "quantity": 1000,
                "unit": "kg",
                "unit_price": 0.5,
                "period": "2026-06",
                "value_status": "sourced",
                "evidence_id": "live_market",
            }
        ],
        financing_cost=20,
        depreciation_cost=30,
        evidence_ids=["live_seed", "live_market"],
        impacts=[
            {
                "category": "water",
                "quantity": 100,
                "unit": "m3",
                "period": "season",
                "value_status": "sourced",
                "evidence_id": "live_water",
            }
        ],
        assumptions=["Yield and price are scenario inputs."],
    )

    assert result["total_cost"] == 250
    assert result["total_revenue"] == 500
    assert result["net_before_financing_tax"] == 300
    assert result["net_after_financing_depreciation_before_tax"] == 250
    assert result["break_even"][0]["break_even_unit_price"] == 0.25
    assert result["break_even"][0]["break_even_quantity"] == 500
    assert result["downside_margin"] == 130
    assert result["cash_flow"][-1]["cumulative_cash_flow"] == 280
    assert result["provenance_status"] == "evidence_linked"
    assert result["evidence_ids"] == ["live_seed", "live_market", "live_water"]

    artifact = store.get_artifact(user["id"], result["artifact_id"])
    workbook = load_workbook(io.BytesIO(storage.get(artifact["storage_path"])))
    assert {
        "Enterprise budget",
        "Break-even",
        "Cash flow",
        "Sensitivity",
        "Sustainability impacts",
        "Assumptions and evidence",
    } == set(workbook.sheetnames)
    assert artifact["metadata"]["downside_margin"] == 130
    assert artifact["metadata"]["evidence_ids"] == result["evidence_ids"]


def test_legacy_budget_inputs_remain_supported(tmp_path) -> None:
    _store, _storage, _user, service = _artifact_service(tmp_path)
    result = service.calculate_enterprise_budget(
        title="Legacy budget",
        currency="USD",
        costs=[{"item": "Seed", "quantity": 1, "unit": "bag", "unit_cost": 10}],
        revenues=[
            {"item": "Produce", "quantity": 2, "unit": "kg", "unit_price": 8}
        ],
        sources=["Farmer estimate"],
    )

    assert result["total_cost"] == 10
    assert result["total_revenue"] == 16
    assert result["net_before_financing_tax"] == 6
    assert result["provenance_status"] == "legacy_unstructured"
    assert result["warning"] == "Scenario estimate; not a profit guarantee."


def test_action_artifact_accepts_evidence_ids_and_legacy_sources(tmp_path) -> None:
    store, _storage, user, service = _artifact_service(tmp_path)
    result = service.generate_farm_action_plan(
        title="Field action",
        context="Observed irrigation issue",
        actions=["Measure soil moisture"],
        evidence_ids=["claim_irrigation_1"],
        sources=["Legacy note"],
    )

    artifact = store.get_artifact(user["id"], result["artifact_id"])
    assert artifact["metadata"]["evidence_ids"] == ["claim_irrigation_1"]
    assert artifact["metadata"]["legacy_sources"] == ["Legacy note"]
    assert artifact["metadata"]["provenance_status"] == "evidence_linked"


@pytest.mark.parametrize(
    ("overrides", "message"),
    [
        ({"as_of_date": "11/08/2026"}, "YYYY-MM-DD"),
        ({"financing_cost": -1}, "finite and non-negative"),
        (
            {
                "costs": [
                    {
                        "item": "Seed",
                        "quantity": 1,
                        "unit": "bag",
                        "unit_cost": 10,
                        "value_status": "sourced",
                    }
                ]
            },
            "requires evidence_id",
        ),
    ],
)
def test_budget_rejects_invalid_material_inputs(tmp_path, overrides, message) -> None:
    _store, _storage, _user, service = _artifact_service(tmp_path)
    arguments = {
        "title": "Invalid budget",
        "currency": "USD",
        "costs": [{"item": "Seed", "quantity": 1, "unit": "bag", "unit_cost": 10}],
        "revenues": [
            {"item": "Produce", "quantity": 2, "unit": "kg", "unit_price": 8}
        ],
    }
    arguments.update(overrides)

    with pytest.raises(ValueError, match=message):
        service.calculate_enterprise_budget(**arguments)


def test_budget_tool_schema_prefers_evidence_and_requires_context() -> None:
    budget = next(
        item["function"]
        for item in ToolRegistry._artifact_definitions()
        if item["function"]["name"] == "calculate_enterprise_budget"
    )
    properties = budget["parameters"]["properties"]
    required = budget["parameters"]["required"]

    assert "evidence_ids" in properties
    assert "sources" in properties
    assert "Legacy" in properties["sources"]["description"]
    assert {"geography", "as_of_date"}.issubset(required)
    assert {"sensitivity_scenarios", "impacts"}.issubset(properties)
