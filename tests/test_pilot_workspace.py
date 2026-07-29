import io
import zipfile

import pytest
from openpyxl import load_workbook

import farmers_chatbot.auth as auth
from farmers_chatbot.artifacts import ArtifactService
from farmers_chatbot.auth import IdentityError, UserIdentity
from farmers_chatbot.documents import DocumentService
from farmers_chatbot.pilot_store import PilotStore, hash_external_identity
from farmers_chatbot.storage_backends import LocalPrivateStorage


def _identity(subject: str, email: str) -> UserIdentity:
    return UserIdentity(
        user_id="",
        issuer="https://accounts.google.com",
        subject=subject,
        email=email,
        name=email.split("@")[0],
        is_admin=False,
    )


def test_verified_google_identity_and_authorization_policies(monkeypatch):
    claims = {
        "iss": "https://accounts.google.com",
        "sub": "google-subject",
        "email": "tester@example.org",
        "email_verified": True,
        "name": "Tester",
    }
    monkeypatch.setattr(auth, "ACCESS_POLICY", "google_any")
    identity = auth.identity_from_claims(claims)
    assert identity.subject == "google-subject"
    assert identity.email == "tester@example.org"

    with pytest.raises(IdentityError):
        auth.identity_from_claims({**claims, "email_verified": False})

    monkeypatch.setattr(auth, "ACCESS_POLICY", "domain_allowlist")
    monkeypatch.setattr(auth, "ALLOWED_DOMAINS", frozenset({"aub.edu.lb"}))
    with pytest.raises(IdentityError):
        auth.identity_from_claims(claims)


def test_workspace_ownership_isolation_and_attachment_cleanup(tmp_path):
    store = PilotStore(sqlite_path=tmp_path / "pilot.sqlite3")
    first = store.upsert_user(_identity("one", "one@example.org"))
    second = store.upsert_user(_identity("two", "two@example.org"))
    project_id = store.create_project(first["id"], "Farm A", "Use field records.")
    conversation_id = store.create_conversation(
        first["id"],
        project_id=project_id,
    )
    attachment_path = (
        f"users/{first['id']}/conversations/{conversation_id}/images/photo.jpg"
    )
    store.add_message(
        first["id"],
        conversation_id,
        role="user",
        content="What does this show?",
        attachments=[{"kind": "image", "storage_path": attachment_path}],
    )

    with pytest.raises(ValueError):
        store.get_project(second["id"], project_id)
    with pytest.raises(ValueError):
        store.list_messages(second["id"], conversation_id)

    deleted_paths = store.delete_conversation(first["id"], conversation_id)
    assert attachment_path in deleted_paths


def test_documents_are_scoped_and_executables_are_rejected(tmp_path):
    store = PilotStore(sqlite_path=tmp_path / "pilot.sqlite3")
    storage = LocalPrivateStorage(tmp_path / "private")
    user = store.upsert_user(_identity("one", "one@example.org"))
    project_id = store.create_project(user["id"], "Potato season")
    documents = DocumentService(store, storage)

    document_id = documents.ingest(
        user["id"],
        project_id,
        filename="field-notes.txt",
        data=b"Soil pH 6.2. Irrigation is recorded weekly.",
        mime_type="text/plain",
    )
    assert store.list_project_chunks(user["id"], project_id)
    assert store.list_documents(user["id"], project_id)[0]["id"] == document_id

    with pytest.raises(ValueError):
        documents.ingest(
            user["id"],
            project_id,
            filename="unsafe.exe",
            data=b"MZ",
            mime_type="application/octet-stream",
        )


def test_docx_and_xlsx_artifacts_open_and_include_controls(tmp_path):
    store = PilotStore(sqlite_path=tmp_path / "pilot.sqlite3")
    storage = LocalPrivateStorage(tmp_path / "private")
    user = store.upsert_user(_identity("one", "one@example.org"))
    service = ArtifactService(store, storage, owner_user_id=user["id"])

    action_plan = service.generate_farm_action_plan(
        title="Irrigation action plan",
        context="A small potato field in Akkar.",
        actions=["Measure soil moisture", "Record irrigation volume"],
        assumptions=["No validated field measurement was supplied."],
        sources=["https://www.fao.org/"],
    )
    action_row = store.get_artifact(user["id"], action_plan["artifact_id"])
    action_bytes = storage.get(action_row["storage_path"])
    assert zipfile.is_zipfile(io.BytesIO(action_bytes))

    calendar = service.generate_crop_calendar(
        title="Potato calendar",
        entries=[
            {
                "period": "=unsafe",
                "activity": "Soil test",
                "trigger": "Before input purchase",
                "risk": "Field-specific timing",
                "source": "FAO",
            }
        ],
        assumptions=["Dates require local validation."],
        sources=["FAO"],
    )
    calendar_row = store.get_artifact(user["id"], calendar["artifact_id"])
    workbook = load_workbook(io.BytesIO(storage.get(calendar_row["storage_path"])))
    assert workbook["Crop calendar"]["A2"].value == "'=unsafe"
    assert "Assumptions and sources" in workbook.sheetnames


def test_whatsapp_identifier_is_hmac_and_not_the_phone_number():
    raw_phone = "96170123456"
    identity_hash = hash_external_identity(raw_phone, "test-secret")
    assert identity_hash != raw_phone
    assert raw_phone not in identity_hash
    assert identity_hash == hash_external_identity(raw_phone, "test-secret")
    assert identity_hash != hash_external_identity(raw_phone, "other-secret")

